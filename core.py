#!/usr/bin/env python
# -*- coding:utf-8 -*-
# author:Administrator
# datetime:2020/2/20 20:21
# software: PyCharm


'''
    读取xlsx表的数据，形成数据库，形成分析报告
    Report类
'''

from tinydb import TinyDB
from tinydb import Query
from openpyxl import load_workbook
from far import _config, _formula

class getDB(object):
    '''
    直接实例化后即可使用
    '''
    def __init__(self, name):
        '''
        getDB类的功能：
            1、获取用户发送的数据文件（xlsx文件）
            2、形成和调用过程数据库（json文件）
            3、形成报告，需要导入(text和para模块)
        属性：
            name：企业名称（xlsx文件名称），被index模块调用，index模块获取name传入
        方法：
            xls_db(): 形成初始数据库
        :param name: 企业名称(xlsx文件名称)
        '''
        self.name = name
        self.initPath()
        self.db = self.xlsx_db()
        self.table_without_nodata = self.db.table('without_nodata')
        self.table_for_print = self.db.table('for_print')
        self.initDB()

    def initPath(self):
        # xlsx文件名
        self.xls_file_name = '%s.xlsx' % self.name
        # xlsx文件根目录
        self.root_path = _config.Path.root + '/safa'
        # xlsx文件路径
        self.xls_file_path = ''.join([self.root_path, r"\input\{}".format(self.xls_file_name)])
        # 读取xlsx文件
        self.ws = load_workbook(
            filename=self.xls_file_path,
            data_only=True)["Sheet1"]

    def initDB(self):
        '''
        还需要对for_print表进行初始化
        :return:
        '''
        self.__remove_zore_year()
        self.calc_f_index('without_nodata')
        self.calc_m_index('without_nodata')
        self.calc_f_index('for_print')
        # self.calc_m_index('for_print')
        self.__remove_zore_item()

    def xlsx_db(self):
        '''
        数据读取/存储
        :param output: xlsx文件数据形成tinydb数据库对象
        :return: tinydb数据库对象
        '''
        self.db_path = ''.join([self.root_path, r"\db\{}.json".format(self.name)])
        db = TinyDB(self.db_path)
        # 清洗数据中因复制粘贴带有的‘\u202c’和数字中含有‘，’的问题
        def clear_data(data):
            if isinstance(data, str):
                data = data.replace('\u202c', '').replace(',', '')
                # ['前3年','前2年','前1年','当期']:
                if data in ['前3年', '前2年', '前1年', '当期']:
                    return data
                else:
                    return float(data)
            else:
                return data
        # 写入数据库
        for row in self.ws.iter_rows(min_row=1,
                                     max_col=self.ws.max_column,
                                     max_row=self.ws.max_row
                                     ):
            for cell in row:
                if cell.value is None or cell.value == ' ' or cell.value == '-':
                    cell.value = 0
            db.insert({
                'items': row[0].value.strip(),
                'year3': clear_data(row[1].value),
                'year2': clear_data(row[2].value),
                'year1': clear_data(row[3].value),
                'month': clear_data(row[4].value),
            })
        return db

    def check_complete(self):
        """
        判断原始数据库db中，没有数据的年份或月份
        :param db: 原始数据库，tinydb类
        :return: 无数据年份列表
        """
        def get_data(db, a, b):
            return db.get(Q.items == a)[b]
        self.db.all()
        Q = Query()
        nodata_years_list = []
        for year in ['year3', 'year2', 'year1', 'month']:
            if get_data(self.db, '资产总计', year) == 0 and \
                    get_data(self.db, '负债合计', year) == 0 and \
                    get_data(self.db, '所有者权益合计', year) == 0:
                nodata_years_list.append(year)
            else:
                pass
        return nodata_years_list

    def __remove_zore_year(self):
        '''
        前置步骤:
            cal_f_index() ---> table('without_nodata')
        清除无数据年份
        row_dict类型为字典，删除无数据年份对应的键值对
        '''
        table_w = self.db.table('without_nodata')
        table_p = self.db.table('for_print')
        row_dict = self.db.all()
        for row in row_dict:
            for year in self.check_complete():
                del row[year]
            table_w.insert(row)
            table_p.insert(row)

    def __remove_zore_item(self):
        '''
        仅对table_for_print使用
        :return:
        '''
        Q = Query()
        self.table_for_print.remove(Q.year3 == 0 and
                                    Q.year2 == 0 and
                                    Q.year1 == 0 and
                                    Q.month == 0)

    def calc_f_index(self, table_name):
        """
        '''纵向扩展 使用insert'''
        1、根据同年数据计算财务指标
        2、存货周转天数、应收账款周转天数 应使用前值与当前值的均值，暂未处理
        3、指标增加需在
            （1）index_list中新增指标名称，
            （2）__formula模块中新增指标计算公式
        :param table: tinydb.table类实例
        :return: tinydb.table类实例
        """
        table = self.db.table(table_name)
        table_ = table.all()
        # Q = Query()
        index_list = [
            '资产负债率', '流动比率', '速动比率', 'EBIT', '利息保障倍数',  # 偿债能力指标
            '营运资产', '营运负债', '营运资金需求', '营运资本', '存货周转天数', '应收账款周转天数',  # 营运能力指标
            '毛利率', '净利润率', '总资产收益率(ROA)', '净资产收益率(ROE)', # 盈利能力指标
            '短债占比', '刚性负债', '刚兑占比', '短期刚兑', '期间费用', '费用收入比', '流动资产占比',  # 报告需要数据
        ]
        years = [
            'year3', 'year2', 'year1', 'month'
        ]
        exsit_data_list = []
        for year in years:
            if year in table_[2].keys():
                exsit_data_list.append(year)

        for index in index_list:
            index_dict = {}
            index_dict.setdefault('items', index)
            for i in exsit_data_list:
                index_dict.setdefault(i, _formula.formula(table, i, index))
            table.insert(index_dict)

        # return table

    def calc_m_index(self, table_name):
        """
            '''横向扩展，使用update'''
            1、增加平均值、求和数、增长率、平均增长率
            2、科目进行分类，更新type标签
        @param table:xlsx转换的tinydb里的table类
        @return: 格式化的数据
        """
        table = self.db.table(table_name)
        table.all()
        Q = Query()

        # 平均值、求和数、增长率、平均增长率
        for row_dict in table.all():
            if row_dict['items'] == '项目':
                pass
            else:
                # 累积值求和
                sum = 0
                exsit_data_list = []
                for year in ['year3', 'year2', 'year1', 'month']:
                    if year in row_dict.keys():
                        sum += row_dict[year]
                        exsit_data_list.append(row_dict[year])
                # 平均值
                averg = sum / len(exsit_data_list)
                # 当期比前值的变化
                if len(exsit_data_list) >= 2:
                    delta = exsit_data_list[-1] - exsit_data_list[-2]
                else:
                    delta = exsit_data_list[-1]
                # 当期比前值的变化幅度
                if exsit_data_list[-2] == 0:
                    ratio = '净新增'
                else:
                    ratio = delta / exsit_data_list[-2]

                table.update(
                    {
                        'averg': averg,
                        'sum': sum,
                        'ratio': ratio,
                        'delta': delta,
                    },
                    Q.items == row_dict['items']
                )
        # 分类标签
        type = [
            '流动资产', '非流动资产', '流动负债', '非流动负债', '权益', '利润表', '现金流量', '财务指标',
        ]  # 数据类型列表
        li_a = [
            '货币资金', '交易性金融资产', '衍生金融资产 ', '应收票据', '应收账款', '应收款项融资 ', '预付账款',
            '其他应收款', '存货', '合同资产', '持有待售资产', '一年到期的非流动资产', '其他流动资产',
        ]  # 流动资产科目
        fi_a = [
            '债权投资', '其他债权投资', '长期应收款', '长期股权投资', '其他权益工具投资', '其他非流动金融资产',
            '投资性房地产', '固定资产', '在建工程', '生产性生物资产', '油气资产', '使用权资产', '无形资产',
            '开发支出', '商誉', '长期待摊费用', '递延所得税资产', '其他非流动资产',
        ]  # 非流动资产科目
        li_d = [
            '短期借款', '交易性金融负债', '衍生金融负债', '应付票据', '应付账款', '预收账款', '合同负债',
            '应付职工薪酬','应交税费', '其他应付款', '持有待售负债', '一年内到期的非流动负债', '其他流动负债',
        ]  # 流动负债科目
        fi_d = [
            '长期借款', '应付债券', '租赁负债', '长期应付款', '预计负债', '递延收益', '递延所得税负债', '其他非流动负债',
        ]  # 非流动负债科目
        equi = [
            '实收资本', '其他权益工具', '资本公积', '其他综合收益', '专项储备', '盈余公积', '未分配利润',
        ]  # 所有者权益科目
        reve = [
            '营业收入', '营业成本', '税金及附加', '销售费用', '管理费用', '研发费用', '财务费用', '投资收益',
            '净敞口套期收益', '公允价值变动收益', '信用减值损失', '资产减值损失', '资产处置收益', '营业利润',
            '营业外收入', '营业外支出', '利润总额', '所得税费用', '净利润',
        ]  # 利润表科目
        csfl = [
            '经营活动现金流入', '销售带来的现金流入', '经营活动现金流出', '经营活动净现金流',
            '投资活动现金流入', '投资活动现金流出', '投资活动现金净流',
            '筹资活动现金流入', '筹资活动现金流出', '筹资活动现金净流入', '现金净流入',
        ]  # 现金流量表科目
        inde = [
            '资产负债率', '流动比率', '速动比率', 'EBIT', '利息保障倍数',
            '营运资产', '营运负债', '营运资金需求', '营运资本', '存货周转天数', '应收账款周转天数',
            '毛利率', '净利润率', '总资产收益率(ROA)', '净资产收益率(ROE)',
        ]  # 财务指标列表
        type_list = [li_a, fi_a, li_d, fi_d, equi, reve, csfl, inde]
        for ty in type_list:
            for it in ty:
                for i in range(len(type_list)):
                    if it in type_list[i]:
                        table.update({'type': type[i]}, Q.items == str(it))
        return table

    def clear_all_data(self):
        self.db.purge()

