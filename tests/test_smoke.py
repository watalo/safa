# -*- coding: utf-8 -*-
"""
safa smoke test
===============
冒烟测试：验证 5 套数据形态都能成功生成 docx 报告。

不验证内容对不对，只验证"能跑通"。

策略：
- 复用 input/模板.xlsx（98 行，59 个有数据科目）
- 5 套数据形态通过列填充/清零来模拟：
    * normal       : 4 列全填（成立 >3 年）
    * no_year3     : 清空第 1 列（成立不足 3 年）
    * no_year2     : 清空前 2 列（成立不足 2 年）
    * no_year1     : 清空前 3 列（成立不足 1 年）
    * all_years    : 4 列都填（这是期末数据场景）
- two_years      : 项目未演示，跳过
"""

import os
import shutil
import sys
import unittest

from openpyxl import load_workbook

# 让 unittest 跑得到 far 包
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from far._config import INPUT_PATH, OUTPUT_PATH, DB_PATH  # noqa: E402
from far.core import getDB  # noqa: E402
from far.main import get_docx  # noqa: E402


# 数据形态定义：每种形态保留右侧 n_keep 列（前面 4-n_keep 列清零）
# 表头是 ['项目', '前3年', '前2年', '前1年', '当期']
# openpyxl 列索引 1-based：1=前3年, 2=前2年, 3=前1年, 4=当期, 0=科目名
# 与 _para.py:115-123 的判断逻辑对齐：'not in type_list' 表示该年没有数据
DATA_TYPE_COLUMNS = {
    'normal':    4,  # 4 列全填：year3/year2/year1/month 都有
    'no_year3':  3,  # 清掉 year3（前 3 年没有），保留 year2/year1/month
    'no_year2':  2,  # 清掉 year3/year2，保留 year1/month
    'no_year1':  1,  # 只留 month（前 3 年 + 前 1 年都没有）
    # 'all_years' 与 'normal' 列数相同但含义不同（month 列要为 0），
    # 由 test_all_years 单独处理
}


def _build_mock_xlsx(name: str, n_keep_right: int) -> str:
    """从 input/模板.xlsx 复制一份，按 n_keep_right 保留右侧几列，左侧清零。"""
    src = os.path.join(INPUT_PATH, '模板.xlsx')
    dst = os.path.join(INPUT_PATH, f'{name}.xlsx')
    shutil.copy(src, dst)

    wb = load_workbook(dst)
    ws = wb.active
    # 清空左侧 (4 - n_keep_right) 列的数据（保持表头不变）
    # 跳过 row[0]（科目名），从列 1（前3年）开始
    n_zero_left = 4 - n_keep_right
    for row in ws.iter_rows(min_row=2):
        for col_idx in range(1, n_zero_left + 1):  # 1-based：列 1 = 前3年
            cell = row[col_idx]  # 0-based 索引: 列 1 = row[1]
            cell.value = 0
    wb.save(dst)
    return dst


def _build_all_years_xlsx(name: str) -> str:
    """all_years 形态：month 列全清 0（year3/year2/year1 都有）"""
    src = os.path.join(INPUT_PATH, '模板.xlsx')
    dst = os.path.join(INPUT_PATH, f'{name}.xlsx')
    shutil.copy(src, dst)

    wb = load_workbook(dst)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        row[4].value = 0  # 第 5 列 = month（openpyxl 索引 4，0-based）
    wb.save(dst)
    return dst


def _cleanup(name: str):
    """清掉 mock 产物：input xlsx（仅 mock 文件，不动模板）, db json, output docx"""
    for path in [INPUT_PATH, OUTPUT_PATH, DB_PATH]:
        for filename in [f'{name}.xlsx', f'{name}.json', f'{name}.docx']:
            full = os.path.join(path, filename)
            if os.path.exists(full):
                try:
                    os.remove(full)
                except OSError:
                    pass


def _print_regression(r: dict, label: str):
    """打印回归测试结果（无论成功失败）"""
    if r['ok']:
        print(f'\n  ✅ {label:<11} docx={r["docx_size"]}B, db_rows={r["db_rows"]}')
    else:
        print(f'\n  ⚠️  {label:<11} 失败: {r.get("err", "unknown")}')


def _run_one(data_type: str) -> dict:
    """跑一种数据形态，返回 {'name', 'xlsx_size', 'docx_size', 'ok', 'err'}"""
    name = f'_smoke_{data_type}'

    result = {'data_type': data_type, 'name': name}

    try:
        # 1) 准备 mock xlsx
        if data_type == 'all_years':
            xlsx_path = _build_all_years_xlsx(name)
        else:
            n_keep_right = DATA_TYPE_COLUMNS[data_type]
            xlsx_path = _build_mock_xlsx(name, n_keep_right)
        result['xlsx_size'] = os.path.getsize(xlsx_path)

        # 2) 跑 getDB（生成 db/<name>.json）
        db = getDB(name=name)
        result['db_rows'] = len(db.table_for_print.all())

        # 3) 跑 get_docx（生成 docx）
        output_path = os.path.join(OUTPUT_PATH, f'{name}.docx')
        get_docx(name=name, output_path=output_path)

        # 4) 验证
        assert os.path.exists(output_path), f'docx 未生成: {output_path}'
        size = os.path.getsize(output_path)
        assert size > 1024, f'docx 太小（{size} 字节），可能内容空'

        result['docx_size'] = size
        result['ok'] = True
        result['err'] = None
    except Exception as e:
        result['ok'] = False
        result['err'] = f'{type(e).__name__}: {e}'
    finally:
        # 5) 清理（避免污染下次测试）
        _cleanup(name)

    return result


class SmokeTest(unittest.TestCase):
    """
    验证 5 套数据形态都能跑通。

    设计原则：
    - 5 套数据形态（normal / no_year3 / no_year2 / no_year1 / all_years）必须全部通过。
    - PR #2 修了 4 个 P0 bug（data_type 命名、financial_sheet、big_change_sheet 括号、webapp 加固）。
    - PR #3 修了 4 个 smoke test regression（format string x2 / s2d4 缺 m 键 / exsit_data_list 越界）。
    - 本 PR (#4) 修了 4 个剩余 bug（big_change_items 列缺失 x2 / {:,.f} 缺精度 / all_years 错用 month 列），
      让 5 套数据形态都达到"能跑通"基线。
    - two_years：项目未演示，跳过。
    """

    def test_normal(self):
        """normal 形态必须能跑通"""
        r = _run_one('normal')
        print(f'\n  ✅ normal:    docx={r.get("docx_size", "?")}B, db_rows={r.get("db_rows", "?")}')
        self.assertTrue(r['ok'], f'normal 失败: {r.get("err")}')

    def test_no_year3(self):
        """no_year3 形态必须能跑通（成立不足 3 年）"""
        r = _run_one('no_year3')
        print(f'\n  ✅ no_year3:  docx={r.get("docx_size", "?")}B, db_rows={r.get("db_rows", "?")}')
        self.assertTrue(r['ok'], f'no_year3 失败: {r.get("err")}')

    def test_no_year2(self):
        """no_year2 形态必须能跑通（成立不足 2 年）"""
        r = _run_one('no_year2')
        print(f'\n  ✅ no_year2:  docx={r.get("docx_size", "?")}B, db_rows={r.get("db_rows", "?")}')
        self.assertTrue(r['ok'], f'no_year2 失败: {r.get("err")}')

    def test_no_year1(self):
        """no_year1 形态必须能跑通（成立不足 1 年）"""
        r = _run_one('no_year1')
        print(f'\n  ✅ no_year1:  docx={r.get("docx_size", "?")}B, db_rows={r.get("db_rows", "?")}')
        self.assertTrue(r['ok'], f'no_year1 失败: {r.get("err")}')

    def test_all_years(self):
        """all_years 形态必须能跑通（期末数据场景，无 month 列）"""
        r = _run_one('all_years')
        print(f'\n  ✅ all_years: docx={r.get("docx_size", "?")}B, db_rows={r.get("db_rows", "?")}')
        self.assertTrue(r['ok'], f'all_years 失败: {r.get("err")}')


if __name__ == '__main__':
    unittest.main(verbosity=2, exit=False)
